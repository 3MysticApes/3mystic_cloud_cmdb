from threemystic_cloud_cmdb.cloud_providers.general.cmdb_connector.base_class.base import cloud_cmdb_general_cmdb_connector_base as base
import urllib
from openpyxl.utils import get_column_letter

class cloud_cmdb_general_cmdb_connector_ms365(base):
  def __init__(self, *args, **kwargs):
    super().__init__(logger_name= "cloud_cmdb_general_cmdb_connector_ms365", *args, **kwargs)
  

  
  def get_cloud_share(self, *args, **kwargs):
    return "ms365"
  
  def _validate_cmdb_init(self, *args, **kwargs):
    self._validate_cmdb_file()
    self._validate_workbook_worksheets()
    self._validate_workbook_worksheets_tables()
    self._validate_workbook_worksheets_tables_columns()

    print(self.get_existing_columns_by_key())
    


    self._get_ms_graph().close_session(session_config = {
      "type":"workbook",
      "drive_id": self.get_cmdb_file().get('id'),
      "persist_changes": True,
      "group_id": self.get_cloud_share_config_value(config_key= self.get_cloud_share()).get('group')
    })
  
  def get_existing_columns_by_key(self, *args, **kwargs):    
    if hasattr(self, "_ms365_existing_columns_by_key"):
      return self._ms365_existing_columns_by_key
    
    self._ms365_existing_columns_by_key = {}
    for container_key, table_data in self._get_worksheet_table_data().items():
      self._ms365_existing_columns_by_key[container_key] = [
        sorted_column.get("name") for sorted_column in sorted(list(table_data.get("extra_columns").get("value").values()), key=lambda x: x.get("index"), reverse= False)
      ]
    
    return self.get_existing_columns_by_key()

  
  def _get_workbook_table_name(self, sheet_name, *args, **kwargs):
    return f'cmdb_{self.get_cmdb_data_containers_display_key()[sheet_name]}'

  def get_cmdb_name(self, *args, **kwargs):
    if hasattr(self, "_cmdb_name_ext"):
      return self._cmdb_name_ext
    
    if not super().get_cmdb_name().endswith(".xlsx"):
      self._cmdb_name_ext = f"{super().get_cmdb_name()}.xlsx"
    
    return self.get_cmdb_name(*args, **kwargs)
  
  def _get_ms_graph_drive_id(self, *args, **kwargs):
    if hasattr(self, "_ms365_graph_drive_id"):
      return self._ms365_graph_drive_id
    
    self._ms365_graph_drive_id = self.get_cloud_share_config_value(
      config_key= self.get_cloud_share()
    )["drive_id"]
    return self._get_ms_graph_drive_id(*args, **kwargs)

  def _get_ms_graph_resource_id(self, *args, **kwargs):
    if hasattr(self, "_ms365_graph_resource_id"):
      return self._ms365_graph_resource_id
    
    self._ms365_graph_resource_id = self.get_cloud_share_config_value(config_key= self.get_cloud_share()).get('group') if self.get_cloud_share_config_value(config_key= self.get_cloud_share()).get('group') != "me" else None
    return self._get_ms_graph_resource_id(*args, **kwargs)

  def _get_ms_graph_resource(self, *args, **kwargs):
    if hasattr(self, "_ms365_graph_resource"):
      return self._ms365_graph_resource
    
    self._ms365_graph_resource = "me" if self.get_cloud_share_config_value(config_key= self.get_cloud_share()).get('group') == "me" else f"groups"
    return self._get_ms_graph_resource(*args, **kwargs)

  def _get_ms_graph(self, *args, **kwargs):
    if hasattr(self, "_ms365_graph"):
      return self._ms365_graph
    
    self._ms365_graph = self.get_common().graph().graph(
      graph_method= "msgraph", 
      credentials= self.get_cloud_client().get_tenant_credential(
        tenant= self.get_cloud_share_config_value(config_key= self.get_cloud_share()).get('tenant_id')))
    
    return self._get_ms_graph(*args, **kwargs)
   
  def get_cmdb_file(self, *args, **kwargs):
    if not hasattr(self, "_ms36_cmdb_file"):
      return None
    
    return self._ms36_cmdb_file
  
  def __set_cmdb_file(self, file_details, *args, **kwargs):
    self._ms36_cmdb_file = file_details

  def _get_ms_graph_base_path(self, drive_item_id, *args, **kwargs):
    return f"{drive_item_id}" if self.get_cloud_share_config_value(config_key= self.get_cloud_share()).get('group') == "me" else f"items/{drive_item_id}"
    
  def _validate_cmdb_file(self, *args, **kwargs):

    local_drive_options = self._get_ms_graph().send_request(
      url = self._get_ms_graph().generate_graph_url(
        resource= self._get_ms_graph_resource(), 
        resource_id= self._get_ms_graph_resource_id(), 
        base_path= f"drive/{self._get_ms_graph_base_path(drive_item_id= self._get_ms_graph_drive_id()[-1].get('id') )}/children")
    )

    if local_drive_options.get("value") != None:
      for item in local_drive_options.get("value"):
        if item.get("name") != self.get_cmdb_name():
          continue

        if item.get("file") is None:
          continue
        

        self.__set_cmdb_file(file_details= item)
        break
    
    if self.get_cmdb_file() is not None:
      return
    
    self._create_cmdb_file()
    self._validate_cmdb_file()

  def _create_cmdb_file(self, *args, **kwargs):
    
    from tempfile import NamedTemporaryFile
    from openpyxl import Workbook
    
    excel_doc = Workbook()
    while len(excel_doc.sheetnames) > 0:
      excel_doc.remove(excel_doc[excel_doc.sheetnames[0]])
    
    for sheet_key, sheet_name in self.get_cmdb_data_containers_key_display().items():
      excel_sheet = excel_doc.create_sheet(sheet_name)
      excel_sheet.append(self.get_cmdb_data_containers_columns().get(sheet_key))
      excel_sheet.freeze_panes = "A2"
    
    
    with NamedTemporaryFile() as tmp:

      excel_doc.save(tmp.name)
      tmp.seek(0)
    
      self._get_ms_graph().send_request(
        url = self._get_ms_graph().generate_graph_url(
          resource= self._get_ms_graph_resource(), 
          resource_id= self._get_ms_graph_resource_id(), 
          base_path= f"drive/{self._get_ms_graph_base_path(drive_item_id= self._get_ms_graph_drive_id()[-1].get('id') )}:/{urllib.parse.quote(self.get_cmdb_name())}:/content"),
          data = tmp.read(),
          params= {"@microsoft.graph.conflictBehavior": "replace"},
          headers= {"Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
          method= "put"
      )
  
  def _get_worksheet_data(self, *args, **kwargs):
    if hasattr(self, "_worksheet_data"):
      return self._worksheet_data
    
    return None

  def _validate_workbook_worksheets(self, *args, **kwargs):

    worksheets_response = self._get_ms_graph().send_request(
      url = self._get_ms_graph().generate_graph_url(
        resource= self._get_ms_graph_resource(), 
        resource_id= self._get_ms_graph_resource_id(), 
        base_path= f"drive/{self._get_ms_graph_base_path(drive_item_id= self.get_cmdb_file().get('id') )}/workbook/worksheets"),
        session_config = {
          "type":"workbook",
          "drive_id": self.get_cmdb_file().get('id'),
          "persist_changes": True,
          "group_id": self.get_cloud_share_config_value(config_key= self.get_cloud_share()).get('group')
        },
        method= "get"
    )

    self._worksheet_data = None
    if worksheets_response.get("value") is None:
      for sheet_name in self.get_cmdb_data_containers_key_display().values():
        self._add_workbook_worksheet(sheet_name= sheet_name)
      return self._validate_workbook_worksheets(*args, **kwargs)
    
    if len(worksheets_response.get("value")) < 1:
      for sheet_name in self.get_cmdb_data_containers_key_display().values():
        self._add_workbook_worksheet(sheet_name= sheet_name)
      return self._validate_workbook_worksheets(*args, **kwargs)
    
    self._worksheet_data = {self.get_cmdb_data_containers_display_key()[worksheet["name"]]:worksheet for worksheet in worksheets_response["value"]}
    missing_worksheets = False
    for data_container_key, data_container_display in self.get_cmdb_data_containers_key_display().items():
      if self._worksheet_data.get(data_container_key) is not None:
        continue

      self._add_workbook_worksheet(sheet_key= data_container_key,sheet_name= data_container_display)
      missing_worksheets = True
    
    if missing_worksheets:      
      return self._validate_workbook_worksheets(*args, **kwargs)
    
    return None

  def _add_workbook_worksheet(self, sheet_key, sheet_name, *args, **kwargs):
    
    self._get_ms_graph().send_request(
      url = self._get_ms_graph().generate_graph_url(
        resource= self._get_ms_graph_resource(), 
        resource_id= self._get_ms_graph_resource_id(), 
        base_path= f"drive/{self._get_ms_graph_base_path(drive_item_id= self.get_cmdb_file().get('id') )}/workbook/worksheets"),
        session_config = {
          "type":"workbook",
          "drive_id": self.get_cmdb_file().get('id'),
          "persist_changes": True,
          "group_id": self.get_cloud_share_config_value(config_key= self.get_cloud_share()).get('group')
        },
        data = {
        "name": sheet_name
        },
        params={"@microsoft.graph.conflictBehavior": "replace"},
        method= "post"
    )

    self.init_workbook_worksheet(sheet_key= sheet_key, sheet_name= sheet_name)
  
  def init_workbook_worksheet(self, sheet_key, sheet_name, *args, **kwargs):
    
    return self._get_ms_graph().send_request(
      url = self._get_ms_graph().generate_graph_url(
        resource= self._get_ms_graph_resource(), 
        resource_id= self._get_ms_graph_resource_id(), 
        base_path= f"drive/{self._get_ms_graph_base_path(drive_item_id= self.get_cmdb_file().get('id') )}/workbook/worksheets/{self._get_worksheet_data()[sheet_key].get('id')}/range(address='A1:{get_column_letter(len(self.get_cmdb_data_containers_columns().get(sheet_name)))}1')"),
        session_config = {
          "type":"workbook",
          "drive_id": self.get_cmdb_file().get('id'),
          "persist_changes": True,
          "group_id": self.get_cloud_share_config_value(config_key= self.get_cloud_share()).get('group')
        },
        data= {
          "values": self.get_cmdb_data_containers_columns().get(sheet_name)
        },
        method= "patch"
    )
  def _get_workbook_worksheets_table_columns(self, sheet_key, table_id, *args, **kwargs):
    columns = self._get_ms_graph().send_request(
      url = self._get_ms_graph().generate_graph_url(
        resource= self._get_ms_graph_resource(), 
        resource_id= self._get_ms_graph_resource_id(), 
        base_path= f"drive/{self._get_ms_graph_base_path(drive_item_id= self.get_cmdb_file().get('id') )}/workbook/worksheets/{self._get_worksheet_data()[sheet_key].get('id')}/tables/{table_id}/columns?$select=id,index,name"),
        method= "get"
    )
    if columns is None:
      return {}
    
    if columns.get("value") is None:
      return []
    
    columns["value"] = {
          column.get("name"):column
          for column in columns["value"]}
    return columns

  def _validate_workbook_worksheets_table(self, sheet_key, sheet_name, table_response, *args, **kwargs):
    if table_response is None:
      return self._add_workbook_worksheet_table(sheet_key= sheet_key, sheet_name= sheet_name)

    if table_response.get("value") is None:
      return self._add_workbook_worksheet_table(sheet_key= sheet_key, sheet_name= sheet_name)

    if len(table_response.get("value")) < 1:
      return self._add_workbook_worksheet_table(sheet_key= sheet_key, sheet_name= sheet_name)
    
    for table in table_response.get("value"):
      if table.get("name") == self._get_workbook_table_name(sheet_name= sheet_name):
        table["extra_columns"] = self._get_workbook_worksheets_table_columns(
          sheet_key= sheet_key,
          table_id= table.get('id')
        )
          
        return table
      
    return self._add_workbook_worksheet_table(sheet_key= sheet_key, sheet_name= sheet_name)

  def _get_worksheet_table_data(self, *args, **kwargs):
    if hasattr(self, "_worksheet_table_data"):
      return self._worksheet_table_data
    
    return None
  
  def _validate_workbook_worksheets_tables(self, *args, **kwargs):
    self._worksheet_table_data = {}
    for sheet_key, sheet_name in self.get_cmdb_data_containers_key_display().items():
      self._worksheet_table_data[sheet_key] = self._validate_workbook_worksheets_table(sheet_key= sheet_key, sheet_name= sheet_name, table_response= self._get_workbook_worksheet_table(sheet_key= sheet_key, sheet_name= sheet_name))

  def _get_workbook_worksheet_table(self, sheet_key, sheet_name, *args, **kwargs):
    return self._get_ms_graph().send_request(
      url = self._get_ms_graph().generate_graph_url(
        resource= self._get_ms_graph_resource(), 
        resource_id= self._get_ms_graph_resource_id(), 
        base_path= f"drive/{self._get_ms_graph_base_path(drive_item_id= self.get_cmdb_file().get('id') )}/workbook/worksheets/{self._get_worksheet_data()[sheet_key].get('id')}/tables"),
        session_config = {
          "type":"workbook",
          "drive_id": self.get_cmdb_file().get('id'),
          "persist_changes": True,
          "group_id": self.get_cloud_share_config_value(config_key= self.get_cloud_share()).get('group')
        },
        method= "get"
    )

  def _add_workbook_worksheet_table(self, sheet_key, sheet_name, *args, **kwargs):
    response = self._get_ms_graph().send_request(
      url = self._get_ms_graph().generate_graph_url(
        resource= self._get_ms_graph_resource(), 
        resource_id= self._get_ms_graph_resource_id(), 
        base_path= f"drive/{self._get_ms_graph_base_path(drive_item_id= self.get_cmdb_file().get('id') )}/workbook/worksheets/{self._get_worksheet_data()[sheet_key].get('id')}/tables/add"),
        session_config = {
          "type":"workbook",
          "drive_id": self.get_cmdb_file().get('id'),
          "persist_changes": True,
          "group_id": self.get_cloud_share_config_value(config_key= self.get_cloud_share()).get('group')
        },
        data = {
        "address": f"{sheet_name}!A1:{get_column_letter(len(self.get_cmdb_data_containers_columns().get(sheet_key)))}1",
        "hasHeaders": True
        },
        params={"@microsoft.graph.conflictBehavior": "replace"},
        method= "post"
    )

    table_id = response.pop("id")
    response["name"] = self._get_workbook_table_name(sheet_name= sheet_name)
    
    return self._get_ms_graph().send_request(
      url = self._get_ms_graph().generate_graph_url(
        resource= self._get_ms_graph_resource(), 
        resource_id= self._get_ms_graph_resource_id(), 
        base_path= f"drive/{self._get_ms_graph_base_path(drive_item_id= self.get_cmdb_file().get('id') )}/workbook/worksheets/{self._get_worksheet_data()[sheet_key].get('id')}/tables/{table_id}"),
        session_config = {
          "type":"workbook",
          "drive_id": self.get_cmdb_file().get('id'),
          "persist_changes": True,
          "group_id": self.get_cloud_share_config_value(config_key= self.get_cloud_share()).get('group')
        },
        data = response,
        params={"@microsoft.graph.conflictBehavior": "replace"},
        method= "patch"
    )

  def _validate_workbook_worksheets_tables_columns(self, *args, **kwargs):
    for sheet_key in self.get_cmdb_data_containers_key_display().keys():
      self._validate_workbook_worksheets_tables_column(sheet_key= sheet_key)
  
  def _add_workbook_worksheets_tables_column(self, sheet_key, column, index, *args, **kwargs):
    return self._get_ms_graph().send_request(
      url = self._get_ms_graph().generate_graph_url(
        resource= self._get_ms_graph_resource(), 
        resource_id= self._get_ms_graph_resource_id(), 
        base_path= f"drive/{self._get_ms_graph_base_path(drive_item_id= self.get_cmdb_file().get('id') )}/workbook/worksheets/{self._get_worksheet_data()[sheet_key].get('id')}/tables/{self._get_worksheet_table_data()[sheet_key].get('id')}/columns/add"),
        session_config = {
          "type":"workbook",
          "drive_id": self.get_cmdb_file().get('id'),
          "persist_changes": True,
          "group_id": self.get_cloud_share_config_value(config_key= self.get_cloud_share()).get('group')
        },
        data = {
          "index": index,
          "name": column
        },
        params={"@microsoft.graph.conflictBehavior": "replace"},
        method= "post"
    )

  def _validate_workbook_worksheets_tables_column(self, sheet_key, *args, **kwargs):
    column_index = -1
    add_columns = []
    
    if self._get_worksheet_table_data()[sheet_key].get("extra_columns") is None:
      self._get_worksheet_table_data()[sheet_key]["extra_columns"] = self._get_workbook_worksheets_table_columns(
          sheet_key= sheet_key,
          table_id= self._get_worksheet_table_data()[sheet_key].get('id')
        )
    
    for column in self.get_cmdb_data_containers_columns_raw_display_byid()[sheet_key].keys():
      column_index += 1

      if column not in self._get_worksheet_table_data()[sheet_key].get("extra_columns").get("value"):        
        add_columns.append({
          "index": column_index,
          "previous_index": column_index - 1,
          "next_index": column_index + 1,
          "column": column
        })
    if len(add_columns) < 1:
      return
  
    for column in add_columns:
      self._add_workbook_worksheets_tables_column(
        sheet_key=sheet_key,
        column= column.get("column"),
        index= column["index"]
      )
    
    self._worksheet_table_data[sheet_key] = self._validate_workbook_worksheets_table(
      sheet_key= sheet_key, sheet_name= self.get_cmdb_data_containers_key_display()[sheet_key], 
      table_response= self._get_workbook_worksheet_table(sheet_key= sheet_key, sheet_name= self.get_cmdb_data_containers_key_display()[sheet_key]))



    # for column in self._get_worksheet_table_data()[sheet_key].get("extra_columns").get("value").items():
    #   print(column)
    # self.get_cmdb_data_containers_columns_raw_display_byid()
    # print(self.get_cmdb_data_containers_columns().get(sheet_key))
    # print(self._get_worksheet_table_data()[sheet_key])

    




    
    
  
